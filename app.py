import os
import cv2
import numpy as np
import pandas as pd
import joblib
from flask import Flask, request, render_template, redirect, url_for, flash
from datetime import date, datetime
from sklearn.neighbors import KNeighborsClassifier
from openpyxl import load_workbook, Workbook
#import mysql.connector


app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Connect to the database
'''
def get_db_connection():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='your_db_password',
        database='users_db'
    )'''


# Global Variables
MESSAGE = "WELCOME! Instruction: To register your attendance, kindly click on 'a' on keyboard."
datetoday = date.today().strftime("%m_%d_%y")
datetoday2 = date.today().strftime("%d-%B-%Y")
face_detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Create folders/files
if not os.path.isdir('Attendance'):
    os.makedirs('Attendance')
if not os.path.isdir('static'):
    os.makedirs('static')
if not os.path.isdir('static/faces'):
    os.makedirs('static/faces')

csv_path = f'Attendance/Attendance-{datetoday}.csv'
excel_path = f'Attendance/Attendance-{datetoday}.xlsx'

if not os.path.isfile(csv_path):
    with open(csv_path, 'w') as f:
        f.write('Name,Roll,Time\n')

if not os.path.isfile(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Roll', 'Time'])
    wb.save(excel_path)

# Helper Functions
def totalreg():
    return len(os.listdir('static/faces'))

def extract_faces(img):
    if img is not None and img.size != 0:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_points = face_detector.detectMultiScale(gray, 1.3, 5)
        return face_points
    return []

def identify_face(facearray):
    model = joblib.load('static/face_recognition_model.pkl')
    return model.predict(facearray)

def train_model():
    faces = []
    labels = []
    userlist = os.listdir('static/faces')
    for user in userlist:
        for imgname in os.listdir(f'static/faces/{user}'):
            img = cv2.imread(f'static/faces/{user}/{imgname}')
            resized_face = cv2.resize(img, (50, 50))
            faces.append(resized_face.ravel())
            labels.append(user)
    knn = KNeighborsClassifier(n_neighbors=5)
    knn.fit(faces, labels)
    joblib.dump(knn, 'static/face_recognition_model.pkl')

def extract_attendance():
    df = pd.read_csv(csv_path)
    return df

def add_attendance(name):
    username = name.split('_')[0]
    userid = name.split('_')[1]
    current_time = datetime.now().strftime("%H:%M:%S")

    df = pd.read_csv(csv_path)
    if not ((df['Name'] == username) & (df['Roll'] == userid)).any():
        with open(csv_path, 'a') as f:
            f.write(f'{username},{userid},{current_time}\n')

        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            ws.append([username, userid, current_time])
            wb.save(excel_path)

def get_registered_users():
    users = os.listdir('static/faces')
    reg_names, reg_rolls, reg_times = [], [], []
    for user in users:
        name, roll = user.split('_')
        reg_names.append(name)
        reg_rolls.append(roll)
        reg_times.append("Registered")  # Placeholder
    return reg_names, reg_rolls, reg_times
from flask import Flask, render_template, request, redirect, url_for, session
import os
import cv2

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Add a secret key for session management

# Hardcoded admin credentials
ADMIN_USERNAME = 'admin123'
ADMIN_PASSWORD = 'password'

@app.route('/')
def home():
    if 'username' not in session:  # Not logged in
        return redirect(url_for('admin_login'))  # Show login page

    # Logged in: show home
    df = extract_attendance()
    today_names = df['Name'].tolist()
    today_rolls = df['Roll'].tolist()
    today_times = df['Time'].tolist()
    today_l = len(today_names)

    reg_names, reg_rolls, reg_times = get_registered_users()

    return render_template('home.html',
                           today_names=today_names,
                           today_rolls=today_rolls,
                           today_times=today_times,
                           today_l=today_l,
                           reg_names=reg_names,
                           reg_rolls=reg_rolls,
                           reg_times=reg_times,
                           reg_l=len(reg_names),
                           totalreg=totalreg(),
                           datetoday2=datetoday2,
                           mess=MESSAGE)

@app.route('/adminlogin', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['username'] = username
            return redirect(url_for('home'))  # Login success: go to home
        else:
            error = 'Invalid credentials! Try again.'

    return render_template('adminlogin.html', error=error)

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['newusername']
        password = request.form['newpassword']
        
        if username not in users_db:  # Check if username already exists
            users_db[username] = {"password": password}  # Add user to the "database"
            session['username'] = username  # Log the user in immediately after signup
            return redirect(url_for('home'))  # Redirect to home after successful signup
        else:
            return render_template('sign.html', error="Username already exists. Please choose another.")

    return render_template('sign.html')

@app.route('/admin.html')
def admin_html():
    return render_template('admin.html')  # Make sure 'admin.html' exists in your templates folder


@app.route('/logout')
def logout():
    session.pop('username', None)  # Remove the username from the session
    return redirect(url_for('admin_login'))  # Redirect to login page after logout

@app.route('/start')
def start():
    global MESSAGE
    ATTENDANCE_MARKED = False

    if not os.path.exists('static/face_recognition_model.pkl'):
        MESSAGE = 'No trained model found. Please add a new face first.'
        return redirect(url_for('home'))

    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, 1.1, 5)

        for (x, y, w, h) in faces:
            face = cv2.resize(frame[y:y+h, x:x+w], (50, 50))
            identified_person = identify_face(face.reshape(1, -1))[0]
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
            cv2.putText(frame, identified_person, (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)

            if cv2.waitKey(1) & 0xFF == ord('a'):
                add_attendance(identified_person)
                MESSAGE = f'Attendance marked for {identified_person}!'
                ATTENDANCE_MARKED = True
                break

        cv2.imshow('Attendance System', frame)

        if ATTENDANCE_MARKED or (cv2.waitKey(1) & 0xFF == ord('q')):
            break

    cap.release()
    cv2.destroyAllWindows()

    return redirect(url_for('home'))

@app.route('/add', methods=['POST'])
def add():
    newusername = request.form['newusername']
    newuserid = request.form['newuserid']
    userimagefolder = f'static/faces/{newusername}_{newuserid}'

    if not os.path.isdir(userimagefolder):
        os.makedirs(userimagefolder)

    cap = cv2.VideoCapture(0)
    i, j = 0, 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        faces = extract_faces(frame)
        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 20), 2)
            if j % 10 == 0:
                name = f'{newusername}_{i}.jpg'
                cv2.imwrite(os.path.join(userimagefolder, name), frame[y:y+h, x:x+w])
                i += 1
            j += 1

        cv2.imshow('Adding New User', frame)

        if cv2.waitKey(1) & 0xFF == ord('q') or i >= 50:
            break

    cap.release()
    cv2.destroyAllWindows()
    train_model()
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True, port=1000)
